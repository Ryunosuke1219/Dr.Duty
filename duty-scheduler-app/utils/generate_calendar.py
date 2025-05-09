
import pandas as pd, openpyxl, datetime, re
from collections import defaultdict
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

def generate_pretty_calendar(schedule_df: pd.DataFrame, year:int, month:int, holidays:set|None=None,
                             out_path:str="pretty_calendar.xlsx")->str:
    if holidays is None:
        holidays=set()
    def to_date(d:int): return datetime.date(year,month,d)
    # collect duty info
    day_info=defaultdict(lambda:{"日直":[],"当直":[],"OC":""})
    for _,row in schedule_df.iterrows():
        day=int(re.match(r'(\d+)',row['Shift']).group(1))
        sub=int(re.search(r'-(\d)',row['Shift']).group(1)) if '-' in row['Shift'] else 0
        names="/".join([n for n in [row['Duty_G0'],row['Duty_G1']] if n])
        if sub==1:
            day_info[day]['日直'].append(names)
        else:
            day_info[day]['当直'].append(names)
            if row['Oncall_G0']:
                day_info[day]['OC']=row['Oncall_G0']
    # workbook
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Calendar"
    # styles
    thin=Side(style="thin",color="000000")
    border=Border(top=thin,left=thin,right=thin,bottom=thin)
    center=Alignment(horizontal="center",vertical="center",wrap_text=True)
    hdr_fill=PatternFill("solid",fgColor="DDDDDD")
    hdr_font=Font(bold=True)
    red_font=Font(color="FF0000")
    blue_font=Font(color="0000FF")
    # title
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    t=ws.cell(1,1,f"{year}年{month}月")
    t.alignment=center; t.font=Font(bold=True,size=14)
    # headers
    weekdays=["日","月","火","水","木","金","土"]
    for col,wd in enumerate(weekdays,1):
        c=ws.cell(2,col,wd); c.alignment=center; c.font=hdr_font; c.fill=hdr_fill; c.border=border
        ws.column_dimensions[get_column_letter(col)].width=22
    for r in range(3,10):
        ws.row_dimensions[r].height=48
    # fill days
    first_wd=to_date(1).weekday() # Mon=0
    col=((first_wd+1)%7)+1
    row=3
    day=1
    while True:
        try:
            date=to_date(day)
        except ValueError:
            break
        info=day_info.get(day,{})
        parts=[]
        if info.get('日直'):
            parts.append("日直 "+"/".join(info['日直']))
        if info.get('当直'):
            line="当直 "+"/".join(info['当直'])
            if info.get('OC'):
                line+="/OC:"+info['OC']
            parts.append(line)
        text=str(day) if not parts else f"{day}\n" + "\n".join(parts)
        cell=ws.cell(row,col,text)
        cell.alignment=center; cell.border=border
        if date in holidays or date.weekday()==6:
            cell.font=red_font
        elif date.weekday()==5:
            cell.font=blue_font
        col+=1
        if col>7:
            col=1; row+=1
        day+=1
    lg=row+2
    ws.merge_cells(start_row=lg,start_column=1,end_row=lg,end_column=7)
    ws.cell(lg,1,"日直 = -1 シフト   当直 = -2 / 平日シフト   OC: Oncall").alignment=center
    wb.save(out_path)
    return out_path
